"""
Optional reader for EBSD post-processing/export workbooks (.xlsx / .xlsm).

These workbooks are produced by commercial EBSD suites (e.g. Oxford AztecCrystal,
Bruker ESPRIT/QUANTAX) as an "Excel export" of an already analysed dataset. They
are NOT raw EBSD maps — they contain *derived* summaries: an Overview metadata
sheet, a Grain List, Boundary Statistics, grain-attribute histograms, pole-figure
profiles and Mackenzie/disorientation distributions.

This module extracts those summaries so the app can:
  * use the workbook Step Size as a calibration value,
  * cross-check grain / boundary / texture statistics,
  * compare its own GND estimate against reference metrics.

Pure openpyxl — no macro execution. The workbook is opened read-only and
`data_only=True`, so only cached cell *values* are read (never formulas or VBA).
"""

from __future__ import annotations

import io
import re
from typing import Any

import numpy as np
import openpyxl


# ──────────────────────────────────────────────────────────────────────────────
#  low-level helpers
# ──────────────────────────────────────────────────────────────────────────────

def _to_float(val: Any) -> float | None:
    """Extract the first numeric value from a cell that may carry a unit string
    such as '0.4405µm', '99.36%', '5.9MB (6,207,983bytes)'."""
    if val is None:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    s = str(val)
    m = re.search(r"-?\d+(?:[.,]\d+)?(?:[eE][+-]?\d+)?", s.replace(",", ""))
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def _sheet_rows(ws) -> list[list]:
    """Materialise a worksheet into a list of row-lists (values only)."""
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _find_sheet(wb, *prefixes: str) -> str | None:
    """Return the first sheet whose name starts with any given prefix
    (case-insensitive). Workbook exporters often truncate sheet names."""
    low = {name.lower(): name for name in wb.sheetnames}
    for pref in prefixes:
        p = pref.lower()
        for lname, real in low.items():
            if lname.startswith(p):
                return real
    return None


def _numeric_stats(values: list[float]) -> dict[str, float] | None:
    arr = np.array([v for v in values if v is not None and np.isfinite(v)], dtype=float)
    if arr.size == 0:
        return None
    return {
        "count": int(arr.size),
        "mean": float(np.mean(arr)),
        "median": float(np.median(arr)),
        "std": float(np.std(arr, ddof=1)) if arr.size > 1 else 0.0,
        "min": float(np.min(arr)),
        "max": float(np.max(arr)),
    }


# ──────────────────────────────────────────────────────────────────────────────
#  section parsers
# ──────────────────────────────────────────────────────────────────────────────

def _parse_overview(ws) -> dict:
    """Overview sheet is a two-column key/value list with section headers.
    Extracts step size (µm), pixel count, raster, hit rate, zero-solution count
    and the phase fraction list."""
    out: dict[str, Any] = {}
    phases: list[dict] = []
    in_phases = False

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        val = row[1] if len(row) > 1 else None
        klow = key.lower()

        # section toggles
        if klow == "phases":
            in_phases = True
            continue
        if klow in ("sample properties", "cleanuphistory", "site notes",
                    "project notes", "specimen notes", "general"):
            in_phases = False

        if in_phases:
            # e.g. "1. Iron bcc (old) (m3m" -> "99.26%"
            m = re.match(r"^\s*(\d+)\.\s*(.+)$", key)
            if m and val is not None:
                phases.append({
                    "index": int(m.group(1)),
                    "name": m.group(2).strip(),
                    "fraction_pct": _to_float(val),
                })
            continue

        if klow.startswith("pixel count"):
            out["pixel_count"] = _to_float(val)
        elif klow.startswith("raster"):
            out["raster"] = str(val).strip() if val is not None else None
            rm = re.match(r"\s*(\d+)\s*[xX]\s*(\d+)", str(val or ""))
            if rm:
                out["raster_cols"] = int(rm.group(1))
                out["raster_rows"] = int(rm.group(2))
        elif klow.startswith("step size"):
            out["step_size_um"] = _to_float(val)
        elif klow.startswith("zero solution count"):
            out["zero_solution_count"] = _to_float(val)
        elif klow.startswith("hit rate"):
            out["hit_rate_pct"] = _to_float(val)
        elif klow.startswith("file name"):
            # keep only the bare file name — never the private absolute path
            raw = str(val or "")
            out["source_file"] = re.split(r"[\\/]", raw)[-1] if raw else None

    if phases:
        out["phases"] = phases
    return out


def _parse_boundary(ws) -> dict:
    """Boundary Statistics: per-phase blocks each containing a '2..10°' (LAGB)
    and a '>10°' (HAGB) row. Returns a per-phase dict with length + fraction."""
    rows = _sheet_rows(ws)
    result: dict[str, dict] = {}
    current_phase = "All Phases"

    for r in rows:
        if not r or all(c is None for c in r):
            continue
        first = str(r[0]).strip() if r[0] is not None else ""
        flow = first.lower()

        # phase header rows carry only a name (no numeric length in col B)
        if first and "°" not in first and flow not in ("boundary",) and _to_float(
            r[1] if len(r) > 1 else None
        ) is None and "µm" not in flow and "%" not in flow:
            current_phase = first
            result.setdefault(current_phase, {})
            continue

        length = _to_float(r[1] if len(r) > 1 else None)
        frac = _to_float(r[-1]) if len(r) > 1 else None
        entry = result.setdefault(current_phase, {})
        if first.startswith("2..10") or "2..10" in first:
            entry["LAGB_length_um"] = length
            entry["LAGB_fraction_pct"] = frac
        elif first.startswith(">10") or ">10" in first:
            entry["HAGB_length_um"] = length
            entry["HAGB_fraction_pct"] = frac

    # drop empty phase entries
    return {k: v for k, v in result.items() if v}


def _parse_grain_list(ws) -> dict:
    """Grain List: row 0 = headers, row 1 = units, row 2+ = data.
    Returns grain count, phase counts and per-attribute numeric stats."""
    rows = _sheet_rows(ws)
    if len(rows) < 3:
        return {}
    headers = [("" if h is None else str(h).strip()) for h in rows[0]]

    # locate columns of interest by header substring
    def col_idx(*subs: str) -> int | None:
        for i, h in enumerate(headers):
            hl = h.lower()
            if any(s in hl for s in subs):
                return i
        return None

    idx = {
        "phase": col_idx("phase"),
        "area": col_idx("area"),
        "ecd": col_idx("equivalent circl"),
        "feret": col_idx("max feret"),
        "mos_mean": col_idx("mean orientation"),
        "mos_max": col_idx("maximum orient"),
        "perimeter": col_idx("perimeter"),
    }

    data = rows[2:]
    grain_count = 0
    phase_counts: dict[str, int] = {}
    cols: dict[str, list] = {k: [] for k in idx if k != "phase" and idx[k] is not None}

    for r in data:
        if not r or r[0] is None:
            continue
        # a valid grain row must have a numeric Id
        if _to_float(r[0]) is None:
            continue
        grain_count += 1
        pi = idx["phase"]
        if pi is not None and pi < len(r) and r[pi] is not None:
            pname = str(r[pi]).strip()
            phase_counts[pname] = phase_counts.get(pname, 0) + 1
        for key, ci in idx.items():
            if key == "phase" or ci is None or key not in cols:
                continue
            if ci < len(r):
                cols[key].append(_to_float(r[ci]))

    stats = {k: _numeric_stats(v) for k, v in cols.items()}
    stats = {k: v for k, v in stats.items() if v is not None}
    return {
        "grain_count": grain_count,
        "phase_counts": phase_counts,
        "stats": stats,
    }


def _parse_pole_sheets(wb) -> list[dict]:
    """Pole-figure profile sheets: columns [Angular Distance (deg), MUD].
    Returns peak / mean MUD and the angle at peak for each such sheet."""
    out = []
    for name in wb.sheetnames:
        if not name.lower().startswith("poleplot"):
            continue
        rows = _sheet_rows(wb[name])
        angles, muds = [], []
        for r in rows[1:]:
            a = _to_float(r[0]) if len(r) > 0 else None
            m = _to_float(r[1]) if len(r) > 1 else None
            if a is not None and m is not None:
                angles.append(a)
                muds.append(m)
        if not muds:
            continue
        muds_a = np.array(muds)
        angles_a = np.array(angles)
        peak_i = int(np.argmax(muds_a))
        out.append({
            "sheet": name,
            "max_mud": float(muds_a.max()),
            "mean_mud": float(muds_a.mean()),
            "angle_at_peak_deg": float(angles_a[peak_i]),
            "n_points": int(muds_a.size),
        })
    return out


def _parse_mackenzie_sheets(wb) -> list[dict]:
    """Mackenzie/disorientation sheets: columns
    [Disorientation Angle, Neighbor Pair, Random Pair, Theoretical].
    Returns the mean measured (neighbor) disorientation and peak angle."""
    out = []
    for name in wb.sheetnames:
        if not name.lower().startswith("mackenzie"):
            continue
        rows = _sheet_rows(wb[name])
        ang, neigh, theo = [], [], []
        for r in rows[1:]:
            a = _to_float(r[0]) if len(r) > 0 else None
            n = _to_float(r[1]) if len(r) > 1 else None
            t = _to_float(r[3]) if len(r) > 3 else None
            if a is None:
                continue
            ang.append(a)
            neigh.append(n if n is not None else np.nan)
            theo.append(t if t is not None else np.nan)
        ang_a = np.array(ang, dtype=float)
        neigh_a = np.array(neigh, dtype=float)
        if ang_a.size == 0 or np.all(np.isnan(neigh_a)):
            continue
        w = np.nan_to_num(neigh_a)
        wsum = w.sum()
        if wsum <= 0:
            # phase present in the file but with no measured neighbour pairs
            continue
        mean_neigh = float((ang_a * w).sum() / wsum)
        peak_i = int(np.nanargmax(neigh_a))
        # LAGB fraction from the measured neighbour distribution (< 15°)
        lagb_mask = ang_a < 15.0
        lagb_frac = float(w[lagb_mask].sum() / wsum)
        out.append({
            "sheet": name,
            "mean_neighbor_disorientation_deg": mean_neigh,
            "peak_angle_deg": float(ang_a[peak_i]),
            "lagb_fraction_lt15deg": lagb_frac,
            "n_bins": int(ang_a.size),
        })
    return out


# ──────────────────────────────────────────────────────────────────────────────
#  public entry point
# ──────────────────────────────────────────────────────────────────────────────

def parse_reference_workbook(file_bytes: bytes) -> dict:
    """
    Parse an EBSD export workbook (.xlsx/.xlsm) and return a dict:

        {
          "overview":  {...},          # step_size_um, pixel_count, raster, hit_rate...
          "boundary":  {phase: {...}}, # LAGB/HAGB lengths & fractions
          "grain":     {...},          # grain_count, phase_counts, per-attr stats
          "pole":      [ {...}, ... ],  # MUD peaks per pole sheet
          "mackenzie": [ {...}, ... ],  # disorientation summaries
          "sheet_names": [...],
        }

    Never raises on partial/unexpected content — missing sections come back empty.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        result: dict[str, Any] = {"sheet_names": list(wb.sheetnames)}

        ov = _find_sheet(wb, "overview")
        result["overview"] = _parse_overview(wb[ov]) if ov else {}

        bd = _find_sheet(wb, "boundary statistic", "boundary")
        result["boundary"] = _parse_boundary(wb[bd]) if bd else {}

        gl = _find_sheet(wb, "grain list")
        result["grain"] = _parse_grain_list(wb[gl]) if gl else {}

        result["pole"] = _parse_pole_sheets(wb)
        result["mackenzie"] = _parse_mackenzie_sheets(wb)
        return result
    finally:
        wb.close()
